# Desafio Aula 10, Módulo 4
import openpyxl


class Bot:
    def __init__(self):
        self.wb = openpyxl.load_workbook('ExemploPlanilha.xlsx')
        self.ws = self.wb['Registros 2015']

    def inicio(self):
        print(self.ws['B11'].value)
        self.ws['B11'] = 'Falcon'
        self.wb.save('ExemploPlanilha.xlsx')
        for linha in self.ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=8):
            for celula in linha:
                print(celula.value)
        for linha in self.ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=8):
            for celula in linha:
                print(celula.value)
        for coluna in self.ws.iter_cols(min_col=3, max_col=3, min_row=2, max_row=None):
            for celula in coluna:
                print(celula.value)


bot = Bot()
bot.inicio()
